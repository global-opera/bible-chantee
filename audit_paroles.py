#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Audit complet des paroles Bible Chantée
Génère un fichier Excel avec toutes les vérifications
"""

import os
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configuration
BASE_DIR = Path(r"C:\ScriptBible\bible-chantee")
LYRICS_DIR = BASE_DIR / "lyrics"
OUTPUT_FILE = BASE_DIR / "AUDIT_PAROLES_COMPLET.xlsx"

# Langues supportées
LANGUAGES = {
    'FR': 'Français',
    'EN': 'English',
    'PT': 'Português',
    'ES': 'Español',
    'DE': 'Deutsch',
    'IT': 'Italiano',
    'TL': 'Tagalog'
}

# Liste des livres (depuis books.js)
BOOKS = [
    {"num": "01", "code": "01_GEN", "name_fr": "Genèse", "chapters": 50},
    {"num": "02", "code": "02_EXO", "name_fr": "Exode", "chapters": 40},
    {"num": "03", "code": "03_LEV", "name_fr": "Lévitique", "chapters": 27},
    {"num": "04", "code": "04_NUM", "name_fr": "Nombres", "chapters": 36},
    {"num": "05", "code": "05_DEU", "name_fr": "Deutéronome", "chapters": 34},
    {"num": "06", "code": "06_JOS", "name_fr": "Josué", "chapters": 24},
    {"num": "07", "code": "07_JDG", "name_fr": "Juges", "chapters": 21},
    {"num": "08", "code": "08_RUT", "name_fr": "Ruth", "chapters": 4},
    {"num": "09", "code": "09_1SAM", "name_fr": "1 Samuel", "chapters": 31},
    {"num": "10", "code": "10_2SAM", "name_fr": "2 Samuel", "chapters": 24},
    {"num": "11", "code": "11_1KI", "name_fr": "1 Rois", "chapters": 22},
    {"num": "12", "code": "12_2KI", "name_fr": "2 Rois", "chapters": 25},
    {"num": "13", "code": "13_1CH", "name_fr": "1 Chroniques", "chapters": 29},
    {"num": "14", "code": "14_2CH", "name_fr": "2 Chroniques", "chapters": 36},
    {"num": "15", "code": "15_EZR", "name_fr": "Esdras", "chapters": 10},
    {"num": "16", "code": "16_NEH", "name_fr": "Néhémie", "chapters": 13},
    {"num": "17", "code": "17_EST", "name_fr": "Esther", "chapters": 10},
    {"num": "18", "code": "18_JOB", "name_fr": "Job", "chapters": 42},
    {"num": "19", "code": "19_PSA", "name_fr": "Psaumes", "chapters": 150},
    {"num": "20", "code": "20_PRO", "name_fr": "Proverbes", "chapters": 31},
    {"num": "21", "code": "21_ECC", "name_fr": "Ecclésiaste", "chapters": 12},
    {"num": "22", "code": "22_SON", "name_fr": "Cantique", "chapters": 8},
    {"num": "23", "code": "23_ISA", "name_fr": "Ésaïe", "chapters": 66},
    {"num": "24", "code": "24_JER", "name_fr": "Jérémie", "chapters": 52},
    {"num": "25", "code": "25_LAM", "name_fr": "Lamentations", "chapters": 5},
    {"num": "26", "code": "26_EZE", "name_fr": "Ézéchiel", "chapters": 48},
    {"num": "27", "code": "27_DAN", "name_fr": "Daniel", "chapters": 12},
    {"num": "28", "code": "28_HOS", "name_fr": "Osée", "chapters": 14},
    {"num": "29", "code": "29_JOE", "name_fr": "Joël", "chapters": 3},
    {"num": "30", "code": "30_AMO", "name_fr": "Amos", "chapters": 9},
    {"num": "31", "code": "31_OBA", "name_fr": "Abdias", "chapters": 1},
    {"num": "32", "code": "32_JON", "name_fr": "Jonas", "chapters": 4},
    {"num": "33", "code": "33_MIC", "name_fr": "Michée", "chapters": 7},
    {"num": "34", "code": "34_NAH", "name_fr": "Nahum", "chapters": 3},
    {"num": "35", "code": "35_HAB", "name_fr": "Habacuc", "chapters": 3},
    {"num": "36", "code": "36_ZEP", "name_fr": "Sophonie", "chapters": 3},
    {"num": "37", "code": "37_HAG", "name_fr": "Aggée", "chapters": 2},
    {"num": "38", "code": "38_ZEC", "name_fr": "Zacharie", "chapters": 14},
    {"num": "39", "code": "39_MAL", "name_fr": "Malachie", "chapters": 4},
    {"num": "40", "code": "40_MAT", "name_fr": "Matthieu", "chapters": 28},
    {"num": "41", "code": "41_MAR", "name_fr": "Marc", "chapters": 16},
    {"num": "42", "code": "42_LUK", "name_fr": "Luc", "chapters": 24},
    {"num": "43", "code": "43_JOH", "name_fr": "Jean", "chapters": 21},
    {"num": "44", "code": "44_ACT", "name_fr": "Actes", "chapters": 28},
    {"num": "45", "code": "45_ROM", "name_fr": "Romains", "chapters": 16},
    {"num": "46", "code": "46_1CO", "name_fr": "1 Corinthiens", "chapters": 16},
    {"num": "47", "code": "47_2CO", "name_fr": "2 Corinthiens", "chapters": 13},
    {"num": "48", "code": "48_GAL", "name_fr": "Galates", "chapters": 6},
    {"num": "49", "code": "49_EPH", "name_fr": "Éphésiens", "chapters": 6},
    {"num": "50", "code": "50_PHP", "name_fr": "Philippiens", "chapters": 4},
    {"num": "51", "code": "51_COL", "name_fr": "Colossiens", "chapters": 4},
    {"num": "52", "code": "52_1TH", "name_fr": "1 Thessaloniciens", "chapters": 5},
    {"num": "53", "code": "53_2TH", "name_fr": "2 Thessaloniciens", "chapters": 3},
    {"num": "54", "code": "54_1TI", "name_fr": "1 Timothée", "chapters": 6},
    {"num": "55", "code": "55_2TI", "name_fr": "2 Timothée", "chapters": 4},
    {"num": "56", "code": "56_TIT", "name_fr": "Tite", "chapters": 3},
    {"num": "57", "code": "57_PHM", "name_fr": "Philémon", "chapters": 1},
    {"num": "58", "code": "58_HEB", "name_fr": "Hébreux", "chapters": 13},
    {"num": "59", "code": "59_JAM", "name_fr": "Jacques", "chapters": 5},
    {"num": "60", "code": "60_1PE", "name_fr": "1 Pierre", "chapters": 5},
    {"num": "61", "code": "61_2PE", "name_fr": "2 Pierre", "chapters": 3},
    {"num": "62", "code": "62_1JN", "name_fr": "1 Jean", "chapters": 5},
    {"num": "63", "code": "63_2JN", "name_fr": "2 Jean", "chapters": 1},
    {"num": "64", "code": "64_3JN", "name_fr": "3 Jean", "chapters": 1},
    {"num": "65", "code": "65_JUD", "name_fr": "Jude", "chapters": 1},
    {"num": "66", "code": "66_REV", "name_fr": "Apocalypse", "chapters": 22}
]

def extract_first_line(lyrics_text):
    """Extrait les 50 premiers caractères de vraies paroles (pas les tags)"""
    if not lyrics_text:
        return ""

    # Nettoyer les tags [Verse 1], [Chorus], etc.
    lines = lyrics_text.split('\n')
    for line in lines:
        line = line.strip()
        # Ignorer les lignes vides et les tags
        if line and not line.startswith('[') and not line.endswith(']'):
            # Retourner les 50 premiers caractères
            return line[:50] + ("..." if len(line) > 50 else "")

    return lyrics_text[:50] if lyrics_text else ""

def check_lyrics_file(lang, book_code, chapter):
    """Vérifie l'existence d'un fichier lyrics et retourne son contenu"""
    # Format: 01_GEN_01_FR.txt
    chapter_str = str(chapter).zfill(2)
    filename = f"{book_code}_{chapter_str}_{lang}.txt"
    filepath = LYRICS_DIR / lang / filename

    if filepath.exists():
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
                # Extraire la section [LYRICS]
                if '[LYRICS]' in content:
                    lyrics = content.split('[LYRICS]')[1].strip()
                else:
                    lyrics = content
                return True, filename, extract_first_line(lyrics)
        except Exception as e:
            return True, filename, f"ERREUR LECTURE: {str(e)}"

    return False, filename, ""

def get_audio_filename(book_code, chapter, lang):
    """Génère le nom du fichier audio attendu"""
    chapter_str = str(chapter).zfill(2)
    return f"{book_code}_{chapter_str}_{lang}.mp3"

def create_worksheet_header(ws):
    """Crée l'en-tête du worksheet avec style"""
    headers = [
        'BookCode', 'BookName', 'Chapter', 'AudioFile', 'LyricsFile',
        'LyricsPreview', 'AudioExists', 'LyricsExists', 'Status', 'Notes'
    ]

    # Style de l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 10  # BookCode
    ws.column_dimensions['B'].width = 20  # BookName
    ws.column_dimensions['C'].width = 8   # Chapter
    ws.column_dimensions['D'].width = 25  # AudioFile
    ws.column_dimensions['E'].width = 25  # LyricsFile
    ws.column_dimensions['F'].width = 50  # LyricsPreview
    ws.column_dimensions['G'].width = 12  # AudioExists
    ws.column_dimensions['H'].width = 12  # LyricsExists
    ws.column_dimensions['I'].width = 12  # Status
    ws.column_dimensions['J'].width = 30  # Notes

def audit_language(lang, wb):
    """Effectue l'audit pour une langue et crée un onglet"""
    print(f"\nAudit de {LANGUAGES[lang]} ({lang})...")

    ws = wb.create_sheet(title=lang)
    create_worksheet_header(ws)

    row = 2
    stats = {'ok': 0, 'error': 0, 'missing': 0}

    for book in BOOKS:
        book_code = book['code']
        book_num = book['num']
        book_name = book['name_fr']

        for chapter in range(1, book['chapters'] + 1):
            # Fichier audio (nom attendu)
            audio_file = get_audio_filename(book_code, chapter, lang)

            # Vérifier fichier lyrics
            lyrics_exists, lyrics_file, lyrics_preview = check_lyrics_file(lang, book_code, chapter)

            # Déterminer le statut
            if lyrics_exists:
                if "ERREUR" in lyrics_preview:
                    status = "ERREUR"
                    stats['error'] += 1
                    notes = "Erreur de lecture du fichier"
                elif not lyrics_preview:
                    status = "ERREUR"
                    stats['error'] += 1
                    notes = "Fichier vide ou format invalide"
                else:
                    status = "OK"
                    stats['ok'] += 1
                    notes = ""
            else:
                status = "MANQUANT"
                stats['missing'] += 1
                notes = "Fichier lyrics introuvable"

            # Écrire la ligne
            ws.cell(row=row, column=1, value=book_num)
            ws.cell(row=row, column=2, value=book_name)
            ws.cell(row=row, column=3, value=chapter)
            ws.cell(row=row, column=4, value=audio_file)
            ws.cell(row=row, column=5, value=lyrics_file)
            ws.cell(row=row, column=6, value=lyrics_preview)
            ws.cell(row=row, column=7, value="À vérifier")  # Audio sur R2
            ws.cell(row=row, column=8, value="OUI" if lyrics_exists else "NON")
            ws.cell(row=row, column=9, value=status)
            ws.cell(row=row, column=10, value=notes)

            # Couleur de statut
            status_cell = ws.cell(row=row, column=9)
            if status == "OK":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == "ERREUR":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            else:  # MANQUANT
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            row += 1

    print(f"   OK: {stats['ok']}")
    print(f"   ERREUR: {stats['error']}")
    print(f"   MANQUANT: {stats['missing']}")

    return stats

def create_summary_sheet(wb, all_stats):
    """Crée une feuille de résumé"""
    ws = wb.create_sheet(title="RÉSUMÉ", index=0)

    # Titre
    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value="AUDIT COMPLET - PAROLES BIBLE CHANTÉE")
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 30

    # En-têtes
    ws.cell(row=3, column=1, value="Langue").font = Font(bold=True)
    ws.cell(row=3, column=2, value="OK").font = Font(bold=True)
    ws.cell(row=3, column=3, value="ERREUR").font = Font(bold=True)
    ws.cell(row=3, column=4, value="MANQUANT").font = Font(bold=True)
    ws.cell(row=3, column=5, value="TOTAL").font = Font(bold=True)

    row = 4
    grand_total = {'ok': 0, 'error': 0, 'missing': 0}

    for lang, stats in all_stats.items():
        total = stats['ok'] + stats['error'] + stats['missing']
        ws.cell(row=row, column=1, value=f"{LANGUAGES[lang]} ({lang})")
        ws.cell(row=row, column=2, value=stats['ok'])
        ws.cell(row=row, column=3, value=stats['error'])
        ws.cell(row=row, column=4, value=stats['missing'])
        ws.cell(row=row, column=5, value=total)

        # Couleurs
        ws.cell(row=row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        grand_total['ok'] += stats['ok']
        grand_total['error'] += stats['error']
        grand_total['missing'] += stats['missing']

        row += 1

    # Total général
    row += 1
    ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=grand_total['ok']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=grand_total['error']).font = Font(bold=True)
    ws.cell(row=row, column=4, value=grand_total['missing']).font = Font(bold=True)
    ws.cell(row=row, column=5, value=sum(grand_total.values())).font = Font(bold=True)

    # Ajuster largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

def main():
    """Fonction principale"""
    print("=" * 60)
    print("AUDIT COMPLET - PAROLES BIBLE CHANTÉE")
    print("=" * 60)

    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # Auditer chaque langue
    all_stats = {}
    for lang in LANGUAGES.keys():
        all_stats[lang] = audit_language(lang, wb)

    # Créer la feuille de résumé
    create_summary_sheet(wb, all_stats)

    # Sauvegarder
    wb.save(OUTPUT_FILE)
    print(f"\nFichier Excel genere: {OUTPUT_FILE}")
    print("\n" + "=" * 60)

    # Rapport texte
    print("\nRAPPORT TEXTUEL:")
    print("=" * 60)
    for lang, stats in all_stats.items():
        total = sum(stats.values())
        print(f"\n{LANGUAGES[lang]} ({lang}):")
        print(f"  OK:        {stats['ok']:4d} ({stats['ok']/total*100:5.1f}%)")
        print(f"  ERREUR:    {stats['error']:4d} ({stats['error']/total*100:5.1f}%)")
        print(f"  MANQUANT:  {stats['missing']:4d} ({stats['missing']/total*100:5.1f}%)")
        print(f"  TOTAL:     {total:4d}")

if __name__ == "__main__":
    main()
